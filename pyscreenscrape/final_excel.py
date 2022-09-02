import openpyxl,os
from main2 import fetch_page_content,screenscrape,dictionary,column2_names,column3_names,column_names
from from_excel import udise_codes

wb = openpyxl.load_workbook("frompython.xlsx")


invalid = []
ws_1 = wb['Basic Details']
ws_2 = wb['Facilities']
ws_3 = wb['Room Details']
ws_4 = wb['Enrolment of Students']
ws_5 = wb['Invalid codes']


for code in udise_codes:
    try:
        f = open("Test.html", "a")
        f.write(fetch_page_content(code))
        f.close()  
    
        screenscrape('Test.html')
        max_row_1 = ws_1.max_row+1
        for i in range(1,20):
            ws_1.cell(row = max_row_1,column=i).value = list(dictionary(column_names).values())[i-1]

        max_row_2 = ws_2.max_row+1
        for i in range(1,22):
            ws_2.cell(row = max_row_2,column=i).value = list(dictionary(column2_names).values())[i-1]

        max_row_3 = ws_3.max_row+1
        for i in range(1,3):
            ws_3.cell(row=max_row_3,column=i).value = list(dictionary(column3_names).values())[i-1]

        max_row_4 = ws_4.max_row+1
        for i in range(1,18):
            ws_4.cell(row=max_row_4,column=i).value = list(screenscrape("Test.html").values())[i-1]

        

        os.remove("Test.html")
        # print(ws_1.max_row)
        # wb.save("frompyy.xlsx")
    except Exception:
        invalid.append(code)
        pass

print(invalid)


for i in range(1,len(invalid)+1):
    ws_5.cell(row=i,column=1).value = invalid[i-1]

wb.save('frompython.xlsx')
