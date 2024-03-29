from openpyxl import Workbook
from main2 import column3_names,column2_names,column_names

header_4 = ['Class', 'Pre-Primary', 'I', 'II', 'III', 'IV', 'V', 'VI', 
'VII', 'VIII', 'IX', 'X', 'XI', 'XII', 'Class(1-12)', 
'Class(1-12) With Pre-Primary', 'Total Teachers']
wb_start = Workbook()
ws_1 = wb_start.active
ws_1.title = "Basic Details"
# ws_1 = wb.create_sheet("Basic Details")
ws_2 = wb_start.create_sheet("Facilities")
ws_3 = wb_start.create_sheet("Room Details")
ws_4 = wb_start.create_sheet("Enrolment of Students")
ws_5 = wb_start.create_sheet('Invalid codes')

try:
    for i in range(1,20):
        ws_1.cell(row=1,column=i).value = column_names[i-1]

    for i in range(1,22):
        ws_2.cell(row=1,column=i).value = column2_names[i-1]

    for i in range(1,3):
        ws_3.cell(row=1,column=i).value = column3_names[i-1]

    for i in range(1,18):
        ws_4.cell(row=1,column=i).value = header_4[i-1]

    wb_start.save("frompython.xlsx")
except Exception:
    pass
